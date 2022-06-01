import multiprocessing
import warnings
import itertools
import statsmodels.api as sm
import matplotlib.pyplot as plt
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime
import requests
from bs4 import BeautifulSoup
from itertools import groupby
import os
import pymysql
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from colorama import init
import pandas as pd
from openpyxl import load_workbook
from termcolor import colored
import multiprocessing as mp
from tkinter import *


init()
sys.setrecursionlimit(10000)


# загрузка патента и упаковка в документ
def patent_loader(url, patent_path, name_process):
    option = Options()
    path = ""
    option.add_argument('--headless')
    driver = webdriver.Chrome(ChromeDriverManager(log_level=0).install(), options=option)
    driver.maximize_window()
    try:
        driver.get(url=url)
        wait = WebDriverWait(driver, 50)
        # сон
        wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class=\'more style-scope classification-viewer\']')))
        full_list = driver.find_element_by_xpath("//div[@class='more style-scope classification-viewer']")
        full_list.click()

        description_patent = driver.find_element_by_xpath('//*[@id="description"]').text

        tables = driver.find_elements_by_xpath('/html/body/search-app/search-result/search-ui/div/div/div/div/div/result-container/patent-result/div/div/div/div[2]/div[1]/section/patent-text/div/section/div/div[*]/tables')
        for table in tables:
            description_patent = description_patent.replace(table.text, "")

        path = patent_path + "/" + name_process + ".html"
        file = open(path, "w", encoding='utf-8')
        file.write(driver.page_source)
        file.close()
    except Exception as _ex:
        print(_ex)
    finally:
        driver.close()
        driver.quit()
    return path, description_patent


# получение того, как был опубликован патент
def get_public_as(id_patent):
    time.sleep(3)
    full_publication=[]
    cookies = {
        '_pk_id.16.72ee': '17f04e5f929a8af6.1648435193.',
        '_pk_ses.16.72ee': '1',
        'SameSite': 'None',
        'ADRUM_BTa': '"R:102|g:1ec43543-8b9a-4f08-9de5-b1593e956b7f|n:epo_68f0c8bd-731c-4814-ab5b-03a81a10e4f3"',
        'ADRUM_BT1': '"R:102|i:1312415|e:12"',
    }

    headers = {
        'Accept': 'application/json,application/i18n+xml',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        # Already added when you pass json=
        # 'Content-Type': 'application/json',
        # Requests sorts cookies= alphabetically
        # 'Cookie': '_pk_id.16.72ee=17f04e5f929a8af6.1648435193.; _pk_ses.16.72ee=1; SameSite=None; ADRUM_BTa="R:102|g:1ec43543-8b9a-4f08-9de5-b1593e956b7f|n:epo_68f0c8bd-731c-4814-ab5b-03a81a10e4f3"; ADRUM_BT1="R:102|i:1312415|e:12"',
        'EPO-Trace-Id': 'u6rdlm-hlogne-AAA-000000',
        'Origin': 'https://worldwide.espacenet.com',
        'Referer': 'https://worldwide.espacenet.com/patent/search/family/057994253/publication/' + id_patent + '?q=' + id_patent,
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
        'X-EPO-PQL-Profile': 'cpci',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="100", "Google Chrome";v="100"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'lang': 'en,de,fr',
        'q': id_patent,
        'qlang': 'cql',
        'p_s': 'espacenet',
        'p_q': 'US10824999B2',
    }

    json_data = {
        'query': {
            'fields': [
                'publications.ti_*',
                'publications.abs_*',
                'publications.pn_docdb',
                'publications.in',
                'publications.inc',
                'publications.pa',
                'publications.pac',
                'publications.pd',
                'publications.pr_docdb',
                'publications.app_fdate.untouched',
                'publications.ipc',
                'publications.ipc_ic',
                'publications.ipc_icci',
                'publications.ipc_iccn',
                'publications.ipc_icai',
                'publications.ipc_ican',
                'publications.ci_cpci',
                'publications.ca_cpci',
                'publications.cl_cpci',
                'biblio:pa;pa_orig;pa_unstd;in;in_orig;in_unstd;pac;inc;pd;pn_docdb;allKindCodes;',
                'oprid_full.untouched',
                'opubd_full.untouched',
            ],
            'from': 0,
            'size': 20,
            'highlighting': [
                {
                    'field': 'publications.ti_en',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.abs_en',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.ti_de',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.abs_de',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.ti_fr',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.abs_fr',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.pn_docdb',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
                {
                    'field': 'publications.pa',
                    'fragment_words_number': 20,
                    'number_of_fragments': 3,
                    'hits_only': True,
                },
            ],
        },
        'filters': {
            'publications.patent': [
                {
                    'value': [
                        'true',
                    ],
                },
            ],
        },
        'widgets': {},
    }

    response = requests.post('https://worldwide.espacenet.com/3.2/rest-services/search', params=params,
                             cookies=cookies, headers=headers, json=json_data)
    json_data = response.json()
    reviews = json_data["hits"][0]["fields"]["biblio"][0]
    for r in reviews:
        pi = json_data["hits"][0]["fields"]["biblio"][0][r]["pn_docdb"]
        l = json_data["hits"][0]["fields"]["biblio"][0][r]["allKindCodes"]
        for i in l:
            full_publication.append(pi[i][0])

    return full_publication


# парсер патнета
def get_patent_bs4(path, desc, publications):
    with open(path, encoding='utf-8') as file:
        src = file.read()
    soup = BeautifulSoup(src, "lxml")
    # start_time = time.monotonic()

    priority_date = ''
    granted_date = ''
    application_items = soup.find("div", class_="wrap style-scope application-timeline").find_all("div", class_="event layout horizontal style-scope application-timeline")
    for application_item in application_items:
        application_text = application_item.find("span", class_="title-text style-scope application-timeline").get_text()
        if "Priority to" in application_text:
            priority_date = application_item.find("div", class_="priority style-scope application-timeline").get_text()
        elif "granted" in application_text:
            granted_date = application_item.find("div", class_="granted style-scope application-timeline").get_text()

    # id патента
    publication = soup.find("div", class_="flex-2 style-scope patent-result").find("h2", id="pubnum").get_text()
    if publication in publications:
        current_datetime = datetime.now()
        time_str = str(current_datetime.hour) + ":" + str(current_datetime.minute) + ":" + str(current_datetime.second)
        # print(colored("["+publication+"]", 'red') + " есть в базе данных ")
        exist_patent = '[' + str(datetime.now().date()) + '][' + time_str + '][' + publication + '] есть в базе данных'
        print(exist_patent)
        patent = {'publication': publication,
                  'priority_date': priority_date,
                  'granted_date': granted_date}
        return 0, exist_patent, patent
    # название патента
    title = soup.find("h1", id="title").text.strip()

    # статус патнета
    status = ''
    status_items = soup.find("div", class_="wrap style-scope application-timeline").find_all("div", class_="event layout horizontal style-scope application-timeline")
    for item in status_items:
        status_text = item.find("div", class_="legal-status style-scope application-timeline")
        if status_text:
            status = item.find("span", class_="title-text style-scope application-timeline").get_text()
            if status != "Active":
                # t = time.monotonic() - start_time
                # print(colored("[" + publication + "]", 'red') + " статус не активен")
                current_datetime = datetime.now()
                time_str = str(current_datetime.hour) + ":" + str(current_datetime.minute) + ":" + str(current_datetime.second)
                non_active_status = '[' + str(datetime.now().date()) + '][' + time_str + '][' + publication + '] статус не активен'
                print(non_active_status)
                patent = {'publication': publication}
                return 1, non_active_status, patent
            break

    # ссылка на патент
    link_origin = "https://patents.google.com/patent/" + publication + "/en"

    # ссылки(а) на ру патенты (доделать)
    links_ru_patent = soup.find_all("state-modifier", class_="style-scope application-timeline")
    links_ru = []
    if len(links_ru_patent):
        for link in links_ru_patent:
            if link.text.strip() == "RU":
                link_ru = "https://patents.google.com/" + link.get("data-result")
                links_ru.append(link_ru.replace('/en', '/ru'))

    # абстракт патента
    abstract = soup.find("div", class_="layout horizontal style-scope patent-text").find("div", class_="abstract style-scope patent-text").text.strip()

    # изобретатели патента(переделать)
    inventors_patent = soup.find("dl", class_="important-people style-scope patent-result").find_all("state-modifier", class_="style-scope patent-result")
    inventors = []
    for inventor in inventors_patent:
        inventors.append(inventor.find("a", id="link").get_text())

    # организация патента(доделать)
    assignee_items = soup.find_all("dd", class_="style-scope patent-result")
    assignee = ""
    for assignee_item in assignee_items:
        if assignee_item.find("a", id="link"):
            pass
        else:
            assignee = assignee_item.text.strip()
            break

    public_as = get_public_as(publication)

    # первый класс патента(доделать)
    classes = []
    full_class = []
    class_patent = soup.find("div", class_="table style-scope classification-viewer").find("div", class_="flex style-scope classification-tree").find_all("span", id="target")
    class_item = class_patent[-1].find("a", id="link").get_text()
    class_desc = class_patent[-1].find("span", class_="description style-scope classification-tree").get_text()
    full_class.append(class_item)
    full_class.append(class_desc)
    classes.append(full_class)

    # остальные классы патента
    classes_patent = soup.find("div", class_="table style-scope classification-viewer").find("div", class_="style-scope classification-viewer").find_all("classification-tree", class_="style-scope classification-viewer")
    for classes_one in classes_patent:
        full_class = []
        classes_item = classes_one.find("div", class_="layout horizontal wrap style-scope classification-tree").find("div", class_="flex style-scope classification-tree").find_all("span", id="target")
        class_item = classes_item[-1].find("a", id="link").get_text()
        class_desc = classes_item[-1].find("span", class_="description style-scope classification-tree").get_text()
        full_class.append(class_item)
        full_class.append(class_desc)
        classes.append(full_class)

    # клеймс патента
    flex_frame = soup.find_all("div", class_="flex flex-width style-scope patent-result")
    claims = flex_frame[1].find("div", class_="layout horizontal style-scope patent-text").find("div", class_="claims style-scope patent-text")
    claims_list = []
    claims_items = claims.find_all("div", class_="claim style-scope patent-text")
    for item in claims_items:
        claims_list.append(item.text.strip())
    no_dup_list = [el for el, _ in groupby(claims_list)]
    claims = "\n".join(no_dup_list)

    # описание патента
    description = desc.replace("Description\n", "", 1)

    patent_citations = []
    family_citations = []
    patent_cited_by = []
    family_cited_by = []
    documents = []

    # цитаты патента
    responsive = soup.find_all("div", class_="responsive-table style-scope patent-result")
    dl = soup.find("dl", class_="links style-scope patent-result")
    dd = dl.find_all("dd", class_="style-scope patent-result")
    i = 0
    for d in dd:
        a = d.find("a").text.strip()
        if "Patent citations" in a:
            # патентные ссылки патента
            citations = responsive[i].find("div", class_="table style-scope patent-result").find("div", class_="tbody style-scope patent-result").find_all("div", class_="tr style-scope patent-result")

            check = False
            for citation in citations:
                full_info = []
                if check is False:
                    check_family = citation.find("a")
                    if check_family:
                        id_cit_p = citation.find("span", class_="td nowrap style-scope patent-result").find(
                            "state-modifier", class_="style-scope patent-result").find("a").get_text()
                        other_param_p = citation.find_all("span", class_="td style-scope patent-result")
                        prior_date_cit_p = other_param_p[0].text.strip()
                        public_date_cit_p = other_param_p[1].text.strip()
                        assignee_cit_p = other_param_p[2].text.strip()
                        title_cit_p = other_param_p[3].text.strip()
                        full_info.append(id_cit_p)
                        full_info.append(prior_date_cit_p)
                        full_info.append(public_date_cit_p)
                        full_info.append(assignee_cit_p)
                        full_info.append(title_cit_p)
                        patent_citations.append(full_info)
                    else:
                        check = True
                        pass
                else:
                    id_cit_f = citation.find("span", class_="td nowrap style-scope patent-result").find(
                        "state-modifier", class_="style-scope patent-result").find("a").get_text()
                    other_param_f = citation.find_all("span", class_="td style-scope patent-result")
                    prior_date_cit_f = other_param_f[0].text.strip()
                    public_date_cit_f = other_param_f[1].text.strip()
                    assignee_cit_f = other_param_f[2].text.strip()
                    title_cit_f = other_param_f[3].text.strip()
                    full_info.append(id_cit_f)
                    full_info.append(prior_date_cit_f)
                    full_info.append(public_date_cit_f)
                    full_info.append(assignee_cit_f)
                    full_info.append(title_cit_f)
                    family_citations.append(full_info)
        elif "Cited by" in a:
            cited_by = responsive[i].find("div", class_="table style-scope patent-result").find("div", class_="tbody style-scope patent-result").find_all("div", class_="tr style-scope patent-result")
            check = False
            for cited in cited_by:
                full_info = []
                if check is False:
                    check_family = cited.find("a")
                    if check_family:
                        id_cited_p = cited.find("span", class_="td nowrap style-scope patent-result").find(
                            "state-modifier", class_="style-scope patent-result").find("a").get_text()
                        other_param_p = cited.find_all("span", class_="td style-scope patent-result")
                        prior_date_cited_p = other_param_p[0].text.strip()
                        public_date_cited_p = other_param_p[1].text.strip()
                        assignee_cited_p = other_param_p[2].text.strip()
                        title_cited_p = other_param_p[3].text.strip()
                        full_info.append(id_cited_p)
                        full_info.append(prior_date_cited_p)
                        full_info.append(public_date_cited_p)
                        full_info.append(assignee_cited_p)
                        full_info.append(title_cited_p)
                        patent_cited_by.append(full_info)
                    else:
                        check = True
                        pass
                else:
                    id_cited_f = cited.find("span", class_="td nowrap style-scope patent-result").find("state-modifier",
                                                                                                       class_="style-scope patent-result").find(
                        "a").get_text()
                    other_param_f = cited.find_all("span", class_="td style-scope patent-result")
                    prior_date_cited_f = other_param_f[0].text.strip()
                    public_date_cited_f = other_param_f[1].text.strip()
                    assignee_cited_f = other_param_f[2].text.strip()
                    title_cited_f = other_param_f[3].text.strip()
                    full_info.append(id_cited_f)
                    full_info.append(prior_date_cited_f)
                    full_info.append(public_date_cited_f)
                    full_info.append(assignee_cited_f)
                    full_info.append(title_cited_f)
                    family_cited_by.append(full_info)
        elif "Similar documents" in a:
            trs = responsive[i].find("div", class_="table style-scope patent-result").find("div", class_="tbody style-scope patent-result").find_all("div", class_="tr style-scope patent-result")
            for tr in trs:
                full_info = []
                other_param_d = tr.find_all("span", class_="td style-scope patent-result")
                id_document = other_param_d[0].text.strip()
                public_date_d = other_param_d[1].text.strip()
                title_d = other_param_d[2].text.strip()
                full_info.append(id_document)
                full_info.append(public_date_d)
                full_info.append(title_d)
                documents.append(full_info)
        elif "Legal events" in a:
            i-=1
        i += 1

    patent = {'publication': publication,
              'title': title,
              'status': status,
              'priority_date': priority_date,
              'granted_date': granted_date,
              'link': link_origin,
              'link_ru': links_ru,
              'abstract': abstract,
              'inventors': inventors,
              'assignee': assignee,
              'public_as': public_as,
              'classes': classes,
              'claims': claims,
              'description': description,
              'patent_citations': patent_citations,
              'family_citations': family_citations,
              'patent_cited_by': patent_cited_by,
              'family_cited_by': family_cited_by,
              'documents': documents}

    # t = time.monotonic() - start_time
    # print(colored("[" + publication + "]", 'blue') + colored('[' + str(round(t, 1)) + ']', 'yellow') + " информация собрана")
    current_datetime = datetime.now()
    time_str = str(current_datetime.hour) + ":" + str(current_datetime.minute) + ":" + str(current_datetime.second)
    full_patent = '[' + str(datetime.now().date()) + '][' + time_str + '][' + publication + '] информация собрана'
    print(full_patent)
    return 2, full_patent, patent


# получение ссылок на патенты
def get_patent_links(cpc):
    patent_links = []
    wb = load_workbook(cpc + "//" + cpc + "_links.xlsx")
    ws = wb.active
    colC = ws['I']

    for row in colC:
        if (row.value == '' or row.value == 'result link' or row.value == None):
            continue
        else:
            url_patent = row.value
            patent_links.append(url_patent)
    return patent_links


# получение документов
def get_doc_patents(cpc):
    headers = {
        'authority': 'patents.google.com',
        'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'x-client-data': 'CJW2yQEIpbbJAQjBtskBCKmdygEIk9vKAQjq8ssBCO/yywEInvnLAQjmhMwBCLWFzAEI/4XMAQjLicwBCI6NzAEYqqnKAQ==',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-dest': 'document',
        'referer': 'https://patents.google.com/xhr/query?url=q%3DCPC%253d' + cpc + '%252f14%252flow%26country%3DUS%26before%3Dpriority%3A20200909%26after%3Dpriority%3A20150101%26status%3DGRANT%26language%3DENGLISH%26type%3DPATENT&exp=&download=true&download_format=xlsx',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'cookie': 'CONSENT=YES+RU.ru+20180311-09-0; ANID=AHWqTUnhydZcNb0Bi9Z2Dol2Fzbfl1sluC1hnyrtfzPNgmk9q2DbO2NGogTpoolK; OGP=-19022591:; __Secure-1PSIDCC=AJi4QfF9uhf29mMsiIaTG7SvO9xaxgz5_i7tJy7kCzsrH1lkOYq8TnvxOwPAweVM-NcPHuXv; HSID=AUf9kWdpkVfcIEwHj; SSID=AHQOR7k5hrQrcWDy2; APISID=rHXMcwYERlmkcJC8/AJ_lEhLnHtSeK12Ee; SAPISID=lneiZkJX8MbzsnpN/AmfmLHsatBHuKK1i_; __Secure-1PAPISID=lneiZkJX8MbzsnpN/AmfmLHsatBHuKK1i_; __Secure-3PAPISID=lneiZkJX8MbzsnpN/AmfmLHsatBHuKK1i_; _ga=GA1.3.1424021396.1631281732; OGPC=19022622-1:; SID=DQgzj7EJRUqoydE9cRjzg5mrgmbgHb0Zo4b6T08xNRiMRKig1NpV0kG5k22OlYKBOHT0cg.; __Secure-1PSID=DQgzj7EJRUqoydE9cRjzg5mrgmbgHb0Zo4b6T08xNRiMRKigLOqnIwPV83QDwfidVBYS8Q.; __Secure-3PSID=DQgzj7EJRUqoydE9cRjzg5mrgmbgHb0Zo4b6T08xNRiMRKigCpoBWLxXnj_r4kRMMohdhw.; S=sso=Pq1sbhI3Cxt7QJFCnYy9FxiGnyuyFlrA; SEARCH_SAMESITE=CgQIg5QB; NID=511=P6MxVTrB07JPZAPoX1ISDORTKkng_1yGNdL6TfmdzbyEuUDchoXgxDB3kM4TMuykPV8VdP4dctB2SKjLrUmmu6Xk1V9CCR-lXhVQwKgpO4f5ZHJ7DFGNv52RziIlr4aT8jXwEEB3N8rUlqXDPk5Bv3jv2AyMSsS9cTvZ9zhTOZdAqfnxIbAmLbUrjLLHKZJeudoT0IkcyIfa2UbNAQbTwdw5MOrFMiXQ-rzeyW9ouXB1UWGi3WCXLpEe70wXn38PKBIImo1PQBS7GaC1qbNL_bDCl68Yt6TcJIEiZoluGiJ2P1BnoOnNenSTbFEVTUsTgpcXG3Y4PatPS6ESXR1Jl--TUmkOCidOA9QHyrwxo6ibMlNUxCW25nke_OYwdedJHdOdpnGlRLg73TQMiybm5MPHL7lCwySAmPPTiDFNVhjvNvEPW0oxhdkzc4Hc4YLw-PZZpAEod9zRGX0qTg; 1P_JAR=2021-11-16-19; SIDCC=AJi4QfGGwRvULEgLIhz7vaDrPkLZva7ep6afDZkFkwBD3MY3vlGf96n9hH-HPJjtC1t3-K_3muhl; __Secure-3PSIDCC=AJi4QfERPdkXEUEQzGAAyXJbehDh1EIb_A-d7zzxvNba8fVbmSlA93b9lcTvSLXDvBUAi4pAmeg',
    }

    params = (
        ('url',
         'q=CPC%3d' + cpc + '%2flow&country=US&before=priority:20200909&after=priority:20150101&status=GRANT&language=ENGLISH&type=PATENT'),
        ('exp', ''),
        ('download', 'true'),
        ('download_format', 'xlsx')
    )
    # time.sleep(5)
    response = requests.get('https://patents.google.com/xhr/query', headers=headers, params=params)
    print(response)
    cpc1 = cpc.replace('/', '-')
    open(cpc1 + "//" + cpc1 + "_links.xlsx", "wb").write(response.content)


def parser_bs4(data):
    name_process = multiprocessing.current_process().name
    result = patent_loader(url=data[0], patent_path="html_for_process", name_process=name_process)
    patent = get_patent_bs4(path=result[0], desc=result[1], publications=data[1])
    if patent[0] == 0:
        parser_bs4.q.put(patent)
        print(colored("[" + patent[2]["publication"] + "]", 'cyan') + " добавлен в очередь")
    elif patent[0] == 2:
        parser_bs4.q.put(patent)
        print(colored("[" + patent[2]["publication"] + "]", 'cyan') + " добавлен в очередь")
    else:
        parser_bs4.q.put(patent)


def parser_bs4_init(q):
    parser_bs4.q = q


# вставить базу патента
def insert_base(publication, status, priority_date,granted_date , title, link, link_ru, assignee, public_as, abstract, claims, description, connection, cursor):
    title = title.replace("\'", "\\'").replace("`", "\\'")
    assignee = assignee.replace("\'", "\\'").replace("`", "\\'")
    abstract = abstract.replace("\'", "\\'").replace("`", "\\'")
    claims = claims.replace("\'", "\\'").replace("`", "\\'")
    description = description.replace("\'", "\\'").replace("`", "\\'")
    public_as1 = ", ".join(public_as)
    if not link_ru:
        link_ru = '-'
    else:
        link_ru = ", ".join(link_ru)
    insert_query = "INSERT INTO `patents` (Publication, Status, Priority_date, Granted_date, Title,Link,Link_ru,Assignee,Public_as,Abstract,Claims,Description) VALUES (" + "\'" + publication + "\'" + "," + "\'" + status + "\'" + "," + "\'" + priority_date + "\'" + "," + "\'" + granted_date + "\'" + ","+ "\'" + title + "\'" + "," + "\'" + link + "\'" + "," + "\'" + link_ru + "\'" + "," + "\'" + assignee + "\'" + "," + "\'" + public_as1 + "\'" + "," + "\'" + abstract + "\'" + "," + "\'" + claims + "\'" + "," + "\'" + description + "\'" + ")" + ";"
    cursor.execute(insert_query)
    connection.commit()


# вставить изобритателей
def insert_inventors(inventors, connection, cursor):
    for inventor in inventors:
        inventor = inventor.replace("\'", "\\'").replace("`", "\\'")
        select_query = "SELECT Name FROM `inventors` WHERE Name = \'" + inventor + "\'"
        cursor.execute(select_query)
        existing_inventor = cursor.fetchone()
        if existing_inventor:
            pass
        else:
            insert_query = "INSERT INTO `inventors` (Name) VALUES (" + "\'" + inventor + "\'" + ")"
            cursor.execute(insert_query)
            connection.commit()


# в ставить классы
def insert_classes(classes, connection, cursor):
    for class_ in classes:
        class_name = class_[0]
        class_descriprion = class_[1].replace("\'", "\\'").replace("`", "\\'")
        select_query = "SELECT Name FROM `classes` WHERE Name = \'" + class_name + "\'"
        cursor.execute(select_query)
        existing_class = cursor.fetchone()
        if existing_class:
            pass
        else:
            insert_query = "INSERT INTO `classes` (Name, Description) VALUES (" + "\'" + class_name + "\'" + "," + "\'" + class_descriprion + "\'" + ")"
            cursor.execute(insert_query)
            connection.commit()


# вставить документы
def insert_documents(documents, connection, cursor):
    for document in documents:
        # document_component = document.split("(razdel)")
        publication = document[0].replace("\'","\\'").replace("`", "\\'")
        date = document[1]
        title = document[2].replace("\'","\\'").replace("`", "\\'")
        select_query = "SELECT Publication FROM `documents` WHERE Publication = " + "\'" + publication + "\'" + ";"
        cursor.execute(select_query)
        existing_document = cursor.fetchone()
        if existing_document:
            pass
        else:
            insert_query = "INSERT INTO `documents` (Publication,Publication_date,Title) VALUES (" + "\'" + publication + "\'" + "," + "\'" + date + "\'"  + "," + "\'" + title + "\'"  + ")" + ";"
            cursor.execute(insert_query)
            connection.commit()


# вставить патентный цитаты
def insert_all_citations(citations, name_table, connection, cursor):
    for citation in citations:
        # citation_component = citation.split("(razdel)")
        publication=citation[0].replace("\'","\\'").replace("`", "\\'")
        priority_date=citation[1]
        publication_date=citation[2]
        assignee=citation[3].replace("\'","\"").replace("`", "\"")
        title=citation[4].replace("\'","\"").replace("`", "\"")
        select_query = "SELECT Publication FROM `"+name_table+"` WHERE Publication = \'" + publication + "\'" + ";"
        cursor.execute(select_query)
        existing_publication = cursor.fetchone()
        if existing_publication:
            pass
        else:
            insert_query = "INSERT INTO `"+name_table+"` (Publication,Priority_date,Publication_date,Assignee,Title) VALUES (" + "\'" + publication + "\'" + "," + "\'" + priority_date + "\'" + "," + "\'" + publication_date + "\'" + "," + "\'" + assignee + "\'" + "," + "\'" + title + "\'" + ")" + ";"
            cursor.execute(insert_query)
            connection.commit()


# вставить изобретателей патенета
def insert_patents_of_inventors(publication, inventors, connection, cursor):
    for inventor in inventors:
        inventor = inventor.replace("\'", "\\'").replace("`", "\\'")
        select_publication_id = "SELECT Id FROM `patents` WHERE publication = \'" + publication + "\'" + ";"
        cursor.execute(select_publication_id)
        publication_id = cursor.fetchone()
        select_inventor_id = "SELECT Id FROM `inventors` WHERE Name = \'" + inventor + "\'" + ";"
        cursor.execute(select_inventor_id)
        inventor_id = cursor.fetchone()
        cursor.execute("SET FOREIGN_KEY_CHECKS=0")
        insert_query = "INSERT INTO `patents_of_inventors` (PatentsId,InventorsId) VALUES (" + "\'" + str(publication_id['Id']) + "\'" + "," + "\'" + str(inventor_id['Id']) + "\'" + ")" + ";"
        cursor.execute(insert_query)
        connection.commit()


# вставить классы патентов
def insert_patents_of_classes(publication, classes, connection, cursor):
    for class_ in classes:
        # class_component = class_.split("(razdel)")
        publication = publication.replace("\'", "\\'").replace("`", "\\'")
        select_publication_id = "SELECT Id FROM `patents` WHERE publication = \'" + publication + "\'" + ";"
        cursor.execute(select_publication_id)
        publication_id = cursor.fetchone()
        select_class_id = "SELECT Id FROM `classes` WHERE Name = \'" + class_[0] + "\'" + ";"
        cursor.execute(select_class_id)
        class_id = cursor.fetchone()
        cursor.execute("SET FOREIGN_KEY_CHECKS=0")
        insert_query = "INSERT INTO `patents_of_classes` (PatentsId,ClassesId) VALUES (" + "\'" + str(publication_id['Id']) + "\'" + "," + "\'" + str(class_id['Id']) + "\'" + ")" + ";"
        cursor.execute(insert_query)
        connection.commit()


# создание многих ко многим
def insert_link_tables(publication, name_sec_table, second_args, name_insert_table, second_id, connection, cursor):
    for second_arg in second_args:
        publication = publication.replace("\'", "\\'").replace("`", "\\'")
        second_arg[0] = second_arg[0].replace("\'", "\\'").replace("`", "\\'")
        select_publication_id = "SELECT Id FROM `patents` WHERE publication = \'" + publication + "\'" + ";"
        cursor.execute(select_publication_id)
        publication_id = cursor.fetchone()
        select_second_arg_id = "SELECT Id FROM `"+name_sec_table+"` WHERE Publication = \'" + second_arg[0] + "\'" + ";"
        cursor.execute(select_second_arg_id)
        second_arg_id = cursor.fetchone()
        cursor.execute("SET FOREIGN_KEY_CHECKS=0")
        insert_query = "INSERT INTO `"+name_insert_table+"` (PatentsId,"+second_id+") VALUES (" + "\'" + str(publication_id['Id']) + "\'" + "," + "\'" + str(second_arg_id['Id']) + "\'" + ")" + ";"
        cursor.execute(insert_query)
        connection.commit()


# обновление базы данных
def update_db(patent, connection, cursor):
    update_dates = "UPDATE patents set Priority_date = \'" + patent['priority_date'] + "\', Granted_date = \'" + patent['granted_date'] + "\' where Publication=\'" + patent['publication'] + "\';"
    cursor.execute(update_dates)
    connection.commit()
    current_datetime = datetime.now()
    time_str = str(current_datetime.hour) + ":" + str(current_datetime.minute) + ":" + str(current_datetime.second)
    full_patent = '[' + str(datetime.now().date()) + '][' + time_str + '][' + patent['publication'] + '] обновлен в базе данных'
    print(full_patent)
    with open("information/db_info.txt", "a+", encoding='utf-8') as file:
        file.write(full_patent + '\n')
        print('[' + patent['publication'] + '] записано в файл [db_info]')


# операции с базой данных
def operations_db(patent, connection, cursor):
    # start_time = time.monotonic()

    # вставляем базу патента
    insert_base(publication=patent['publication'],
                status=patent['status'],
                priority_date=patent['priority_date'],
                granted_date=patent['granted_date'],
                title=patent['title'],
                link=patent['link'],
                link_ru=patent['link_ru'],
                assignee=patent['assignee'],
                public_as=patent['public_as'],
                abstract=patent['abstract'],
                claims=patent['claims'],
                description=patent['description'],
                connection=connection,
                cursor=cursor)

    insert_inventors(inventors=patent['inventors'], connection=connection, cursor=cursor)

    insert_classes(classes=patent['classes'], connection=connection, cursor=cursor)

    insert_documents(documents=patent['documents'], connection=connection, cursor=cursor)

    insert_all_citations(citations=patent['patent_citations'], name_table='patent_citations', connection=connection, cursor=cursor)
    insert_all_citations(citations=patent['family_citations'], name_table='family_citations', connection=connection, cursor=cursor)
    insert_all_citations(citations=patent['patent_cited_by'], name_table='patent_cited_by', connection=connection, cursor=cursor)
    insert_all_citations(citations=patent['family_cited_by'], name_table='family_cited_by', connection=connection, cursor=cursor)

    insert_patents_of_inventors(publication=patent['publication'], inventors=patent['inventors'], connection=connection, cursor=cursor)
    insert_patents_of_classes(publication=patent['publication'], classes=patent['classes'], connection=connection, cursor=cursor)

    insert_link_tables(publication=patent['publication'],
                       name_sec_table='patent_citations',
                       second_args=patent['patent_citations'],
                       name_insert_table='patents_of_pc',
                       second_id='PcId',
                       connection=connection, cursor=cursor)

    insert_link_tables(publication=patent['publication'],
                       name_sec_table='family_citations',
                       second_args=patent['family_citations'],
                       name_insert_table='patents_of_fc',
                       second_id='FcId',
                       connection=connection, cursor=cursor)

    insert_link_tables(publication=patent['publication'],
                       name_sec_table='patent_cited_by',
                       second_args=patent['patent_cited_by'],
                       name_insert_table='patents_of_pcb',
                       second_id='PcbId',
                       connection=connection, cursor=cursor)

    insert_link_tables(publication=patent['publication'],
                       name_sec_table='family_cited_by',
                       second_args=patent['family_cited_by'],
                       name_insert_table='patents_of_fcb',
                       second_id='FcbId',
                       connection=connection, cursor=cursor)

    insert_link_tables(publication=patent['publication'],
                       name_sec_table='documents',
                       second_args=patent['documents'],
                       name_insert_table='patents_of_documents',
                       second_id='DocumentsId',
                       connection=connection, cursor=cursor)
    # t = time.monotonic() - start_time
    # print(colored("[" + patent["publication"] + "]", 'green') + colored('[' + str(round(t, 1)) + ']', 'yellow') + " добавлен в базу данных")
    current_datetime = datetime.now()
    time_str = str(current_datetime.hour) + ":" + str(current_datetime.minute) + ":" + str(current_datetime.second)
    full_patent = '[' + str(datetime.now().date()) + '][' + time_str + '][' + patent['publication'] + '] добавлен в базу данных'
    print(full_patent)
    with open("information/db_info.txt", "a+", encoding='utf-8') as file:
        file.write(full_patent + '\n')
        print('[' + patent['publication'] + '] записано в файл [db_info]')


# временно
def insert_patents_db(q, len_patents):
    connection = pymysql.connect(
        host="localhost",
        port=3306,
        user="monty",
        password="some_pass",
        database="Patents",
        cursorclass=pymysql.cursors.DictCursor
    )
    cursor = connection.cursor()
    l = len_patents
    i=0
    while True:
        if len_patents != 0:
            if q.empty():
                print('очередь спит')
                time.sleep(2)
            else:
                patent = q.get()
                if patent is None:
                    pass
                else:

                    # запись в бд
                    operations_db(patent=patent, connection=connection, cursor=cursor)
                i+=1
                print("прогресс ["+str(i)+"/"+str(l)+"]")
                len_patents = len_patents-1
        else:
            cursor.close()
            break


# подключение к базе данных
def connect_db():
    print("подключение к базе данных ", end='')
    try:
        connection = pymysql.connect(
            host="localhost",
            port=3306,
            user="monty",
            password="some_pass",
            database="Patents",
            cursorclass=pymysql.cursors.DictCursor
        )
        print(colored("[успешно]", 'green'))
        return connection
    except Exception as ex:
        print(colored("[безуспешно]", 'red'))
        print(ex)


# создание ссылок
def create_links_patents(cpc_classes):
    all_patent_links = []
    for cpc in cpc_classes:
        cpc1 = cpc.replace('/','-')
        if not os.path.isdir(cpc1):
            os.mkdir(cpc1)
            print("Создание директории [" + cpc1 + "]")
        if os.path.isfile(cpc1 + "/" + cpc1 + "_links.xlsx"):
            patent_links = get_patent_links(cpc=cpc1)
        else:
            get_doc_patents(cpc=cpc)
            patent_links = get_patent_links(cpc=cpc1)
        all_patent_links.extend(patent_links)
        print("cpc: " + "[" + cpc + "]")
        print("Кол-во патентов: " + "[" + str(len(patent_links)) + "]")
    print("Общее кол-во ссылок: " + str(len(all_patent_links)))
    return all_patent_links


def main():
    print("-----------------------------------------------+")
    print("| ВЫБЕРИТЕ ФУНКЦИЮ                             |")
    print("+----------------------------------------------+")
    print("| " + colored("1) ", 'green') + "ПАРСЕР                                    |\n" +
          "| " + colored("2) ", 'green') + "АНАЛИЗ                                    |")
    print("+----------------------------------------------+\n")

    print(colored("◆", 'green') + " - " + "важные функции")
    print(colored("◆", 'blue') + " - " + "побочные функции")
    print(colored("◆", 'red') + " - " + "не работает")

    # Подключение к базе данных
    connection = connect_db()
    cursor = connection.cursor()

    # Получение списка патентов из бд
    exist_publications = []
    exist_publication = "SELECT Publication FROM patents;"
    cursor.execute(exist_publication)
    exist_publication = cursor.fetchall()
    for publication in exist_publication:
        exist_publications.append(publication["Publication"])


    while True:
        print("\n" + "ВВЕДИТЕ НОМЕР: ", end='')
        num = input()

        if num == "1":
            start_time = time.monotonic()
            cpc_classes = ["H04N5", "H04N7", "H04N5/2258"]
            all_patent_links = create_links_patents(cpc_classes=cpc_classes)
            if not os.path.isdir("html_for_process"):
                os.mkdir("html_for_process")
                print("Создание директории [html_for_process]")
            if not os.path.isdir("information"):
                os.mkdir("information")
                print("Создание директории [information]")
            q = mp.Queue()
            p = mp.Pool(3, parser_bs4_init, [q])
            p.imap(parser_bs4, ((elt, exist_publications) for elt in all_patent_links))
            p.close()

            count_all_patents = len(all_patent_links) - 1
            num_patent = 0
            while True:
                if count_all_patents != 0:
                    if q.empty():
                        # print(q.qsize())
                        print('очередь пуста')
                        time.sleep(5)
                    else:
                        patent = q.get()
                        if patent[0] == 0:
                            update_db(patent=patent[2], connection=connection, cursor=cursor)
                            with open("information/patents_info.txt", "a+", encoding='utf-8') as file:
                                file.write(patent[1] + '\n')
                                print('[' + patent[2]['publication'] + '] записано в файл [patents_info]')
                        elif patent[0] == 1:
                            with open("information/patents_info.txt", "a+", encoding='utf-8') as file:
                                file.write(patent[1] + '\n')
                                print('[' + patent[2]['publication'] + '] записано в файл [patents_info]')
                        elif patent[0] == 2:
                            operations_db(patent=patent[2], connection=connection, cursor=cursor)
                        num_patent += 1
                        print("прогресс [" + str(num_patent) + "/" + str(len(all_patent_links)) + "]")
                        count_all_patents = count_all_patents - 1
                else:
                    cursor.close()
                    break
            t = time.monotonic() - start_time
            print('парсинг окончен [' + str(t) + ']')
        elif num == "2":
            cpc_classes = ["H04N5/2258"]
            all_patent_links = create_links_patents(cpc_classes=cpc_classes)
            none_list=[]
            f_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            s_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            sev_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            e_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            n_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            t_year = [[],[],[],[], [],[],[],[], [],[],[],[],]
            for link in all_patent_links:
                link = link.replace('/en', '').replace('https://patents.google.com/patent/', '')
                select_publication_id = "SELECT Publication, Priority_date FROM patents WHERE Publication = \'"+link+"\';"
                cursor.execute(select_publication_id)
                i = cursor.fetchone()
                connection.commit()

                if i is None:
                    none_list.append(link)
                else:
                    year = i['Priority_date'][0:4]
                    if year == "2015":
                        if "-01-" in i['Priority_date']:
                            f_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            f_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            f_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            f_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            f_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            f_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            f_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            f_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            f_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            f_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            f_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            f_year[11].append(i)
                    elif year == "2016":
                        if "-01-" in i['Priority_date']:
                            s_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            s_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            s_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            s_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            s_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            s_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            s_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            s_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            s_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            s_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            s_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            s_year[11].append(i)
                    elif year == "2017":
                        if "-01-" in i['Priority_date']:
                            sev_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            sev_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            sev_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            sev_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            sev_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            sev_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            sev_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            sev_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            sev_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            sev_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            sev_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            sev_year[11].append(i)
                    elif year == "2018":
                        if "-01-" in i['Priority_date']:
                            e_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            e_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            e_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            e_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            e_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            e_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            e_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            e_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            e_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            e_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            e_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            e_year[11].append(i)
                    elif year == "2019":
                        if "-01-" in i['Priority_date']:
                            n_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            n_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            n_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            n_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            n_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            n_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            n_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            n_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            n_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            n_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            n_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            n_year[11].append(i)
                    elif year == "2020":
                        if "-01-" in i['Priority_date']:
                            t_year[0].append(i)
                        elif "-02-" in i['Priority_date']:
                            t_year[1].append(i)
                        elif "-03-" in i['Priority_date']:
                            t_year[2].append(i)
                        elif "-04-" in i['Priority_date']:
                            t_year[3].append(i)
                        elif "-05-" in i['Priority_date']:
                            t_year[4].append(i)
                        elif "-06-" in i['Priority_date']:
                            t_year[5].append(i)
                        elif "-07-" in i['Priority_date']:
                            t_year[6].append(i)
                        elif "-08-" in i['Priority_date']:
                            t_year[7].append(i)
                        elif "-09-" in i['Priority_date']:
                            t_year[8].append(i)
                        elif "-10-" in i['Priority_date']:
                            t_year[9].append(i)
                        elif "-11-" in i['Priority_date']:
                            t_year[10].append(i)
                        elif "-12-" in i['Priority_date']:
                            t_year[11].append(i)



            x2 = ["2015-01", "2015-02", "2015-03", "2015-04", "2015-05", "2015-06", "2015-07", "2015-08", "2015-09", "2015-10", "2015-11", "2015-12",
                  "2016-01", "2016-02", "2016-03", "2016-04", "2016-05", "2016-06", "2016-07", "2016-08", "2016-09", "2016-10", "2016-11", "2016-12",
                  "2017-01", "2017-02", "2017-03", "2017-04", "2017-05", "2017-06", "2017-07", "2017-08", "2017-09", "2017-10", "2017-11", "2017-12",
                  "2018-01", "2018-02", "2018-03", "2018-04", "2018-05", "2018-06", "2018-07", "2018-08", "2018-09", "2018-10", "2018-11", "2018-12",
                  "2019-01", "2019-02", "2019-03", "2019-04", "2019-05", "2019-06", "2019-07", "2019-08", "2019-09", "2019-10", "2019-11", "2019-12"]
            x1=[]
            for i in x2:
                da = pd.Timestamp(i)
                x1.append(da)



            y = [len(f_year[0]), len(f_year[1]), len(f_year[2]), len(f_year[3]),len(f_year[4]), len(f_year[5]), len(f_year[6]), len(f_year[7]),len(f_year[8]), len(f_year[9]), len(f_year[10]), len(f_year[11]),
                 len(s_year[0]), len(s_year[1]), len(s_year[2]), len(s_year[3]), len(s_year[4]), len(s_year[5]), len(s_year[6]), len(s_year[7]), len(s_year[8]), len(s_year[9]), len(s_year[10]), len(s_year[11]),
                 len(sev_year[0]), len(sev_year[1]), len(sev_year[2]), len(sev_year[3]), len(sev_year[4]), len(sev_year[5]), len(sev_year[6]), len(sev_year[7]), len(sev_year[8]), len(sev_year[9]), len(sev_year[10]), len(sev_year[11]),
                 len(e_year[0]), len(e_year[1]), len(e_year[2]), len(e_year[3]), len(e_year[4]), len(e_year[5]), len(e_year[6]), len(e_year[7]), len(e_year[8]), len(e_year[9]), len(e_year[10]), len(e_year[11]),
                 len(n_year[0]), len(n_year[1]), len(n_year[2]), len(n_year[3]), len(n_year[4]), len(n_year[5]), len(n_year[6]), len(n_year[7]), len(n_year[8]), len(n_year[9]), len(n_year[10]), len(n_year[11])]
            b = pd.Series(y, index = x1)
            print(b)

            b.plot(figsize=(15, 6))
            plt.show()

            # Определите p, d и q в диапазоне 0-2
            p = d = q = range(0, 2)
            # Сгенерируйте различные комбинации p, q и q
            pdq = list(itertools.product(p, d, q))
            # Сгенерируйте комбинации сезонных параметров p, q и q
            seasonal_pdq = [(x[0], x[1], x[2], 12) for x in list(itertools.product(p, d, q))]
            print('Examples of parameter combinations for Seasonal ARIMA...')
            print('SARIMAX: {} x {}'.format(pdq[1], seasonal_pdq[1]))
            print('SARIMAX: {} x {}'.format(pdq[1], seasonal_pdq[2]))
            print('SARIMAX: {} x {}'.format(pdq[2], seasonal_pdq[3]))
            print('SARIMAX: {} x {}'.format(pdq[2], seasonal_pdq[4]))
            re1 = []
            re3 = []
            warnings.filterwarnings("ignore")  # отключает предупреждения
            for param in pdq:
                for param_seasonal in seasonal_pdq:
                    re2=[]
                    try:
                        mod = sm.tsa.statespace.SARIMAX(b,
                                                        order=param,
                                                        seasonal_order=param_seasonal,
                                                        enforce_stationarity=False,
                                                        enforce_invertibility=False)
                        results = mod.fit()
                        print('ARIMA{}x{}12 - AIC:{}'.format(param, param_seasonal, results.aic))
                        re2.append(param)
                        re2.append(param_seasonal)
                        re2.append(results.aic)
                        re1.append(re2)
                        re3.append(results.aic)
                    except:
                        continue

            o = min(re3)

            for item in re1:
                if o == item[2]:
                    opi = item
                    print(item)
                    break
            print(min(re3))
            # ARIMA(1, 1, 1)x(1, 1, 1, 4)4 - AIC:72.89707964599648
            mod = sm.tsa.statespace.SARIMAX(b,
                                            order=opi[0],
                                            seasonal_order=opi[1],
                                            enforce_invertibility=False)
            results = mod.fit()
            print(results.summary().tables[1])

            results.plot_diagnostics(figsize=(15, 12))
            plt.show()

            pred = results.get_prediction(start=pd.to_datetime('2018-01-01'), dynamic=False)
            pred_ci = pred.conf_int()

            ax = b['2015':].plot(label='observed')
            pred.predicted_mean.plot(ax=ax, label='One-step ahead Forecast', alpha=.7)
            ax.fill_between(pred_ci.index,
                            pred_ci.iloc[:, 0],
                            pred_ci.iloc[:, 1], color='k', alpha=.2)
            ax.set_xlabel('Date')
            ax.set_ylabel('CO2 Levels')
            plt.legend()
            plt.show()

            y_forecasted = pred.predicted_mean
            y_truth = b['2017-01-01':]
            # Compute the mean square error
            mse = ((y_forecasted - y_truth) ** 2).mean()
            print('The Mean Squared Error of our forecasts is {}'.format(round(mse, 2)))

            pred_dynamic = results.get_prediction(start=pd.to_datetime('2017-01-01'), dynamic=True, full_results=True)
            pred_dynamic_ci = pred_dynamic.conf_int()

            ax = b['2015':].plot(label='observed', figsize=(20, 15))
            pred_dynamic.predicted_mean.plot(label='Dynamic Forecast', ax=ax)
            ax.fill_between(pred_dynamic_ci.index,
                            pred_dynamic_ci.iloc[:, 0],
                            pred_dynamic_ci.iloc[:, 1], color='k', alpha=.25)
            ax.fill_betweenx(ax.get_ylim(), pd.to_datetime('2017-01-01'), b.index[-1],
                             alpha=.1, zorder=-1)
            ax.set_xlabel('Date')
            ax.set_ylabel('CO2 Levels')
            plt.legend()
            plt.show()

            # Извлечь прогнозируемые и истинные значения временного ряда
            y_forecasted = pred_dynamic.predicted_mean
            y_truth = b['2017-01-01':]
            # Вычислить среднеквадратичную ошибку
            mse = ((y_forecasted - y_truth) ** 2).mean()
            print('The Mean Squared Error of our forecasts is {}'.format(round(mse, 2)))

            # Получить прогноз на 500 шагов вперёд
            pred_uc = results.get_forecast(steps=12)
            # Получить интервал прогноза
            pred_ci = pred_uc.conf_int()

            ax = b.plot(label='observed', figsize=(20, 15))
            pred_uc.predicted_mean.plot(ax=ax, label='Forecast')
            ax.fill_between(pred_ci.index,
                            pred_ci.iloc[:, 0],
                            pred_ci.iloc[:, 1], color='k', alpha=.25)
            ax.set_xlabel('Date')
            ax.set_ylabel('CO2 Levels')
            plt.legend()
            plt.show()
        else:
            print(colored("НЕИЗВЕСТНО", 'red'))


if __name__ == "__main__":
    main()




